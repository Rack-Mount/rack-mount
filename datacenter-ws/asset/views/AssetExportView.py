import io
import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiResponse
from drf_spectacular.types import OpenApiTypes

from asset.models import Asset
from asset.views.AssetViewSet import AssetFilter


COLUMNS = [
    ('Hostname', lambda a: a.hostname or ''),
    ('Modello', lambda a: a.model.name if a.model else ''),
    ('Rack Units', lambda a: a.model.rack_units if a.model else ''),
    ('Vendor', lambda a: a.model.vendor.name if a.model and a.model.vendor else ''),
    ('Tipo', lambda a: a.model.type.name if a.model and a.model.type else ''),
    ('Stato', lambda a: a.state.name if a.state else ''),
    ('Seriale', lambda a: a.serial_number or ''),
    ('SAP ID', lambda a: a.sap_id or ''),
    ('Order ID', lambda a: a.order_id or ''),
    ('Alimentatori', lambda a: a.power_supplies),
    ('Assorbimento (W)', lambda a: a.power_cosumption_watt),
    ('Note', lambda a: a.note or ''),
    ('Creato', lambda a: a.created_at.strftime(
        '%d/%m/%Y') if a.created_at else ''),
    ('Aggiornato', lambda a: a.updated_at.strftime(
        '%d/%m/%Y') if a.updated_at else ''),
    ('Scad. garanzia', lambda a: a.warranty_expiration.strftime(
        '%d/%m/%Y') if a.warranty_expiration else ''),
    ('Scad. supporto', lambda a: a.support_expiration.strftime(
        '%d/%m/%Y') if a.support_expiration else ''),
    ('Data acquisto', lambda a: a.purchase_date.strftime(
        '%d/%m/%Y') if a.purchase_date else ''),
    ('Data dismissione', lambda a: a.decommissioned_date.strftime(
        '%d/%m/%Y') if a.decommissioned_date else ''),
]

HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
HEADER_BORDER = Border(
    bottom=Side(style='thin', color='2F5496'),
)
ROW_FONT = Font(color='000000', size=10)
ROW_FILL_ODD = PatternFill('solid', fgColor='FFFFFF')
ROW_FILL_EVEN = PatternFill('solid', fgColor='DCE6F1')


class AssetExportView(APIView):
    """
    GET /asset/export
    Accepts the same filters as AssetViewSet (search, state, model__type, ordering, ids).
    Returns an .xlsx file.
    """
    filter_backends = (filters.OrderingFilter,
                       filters.SearchFilter, DjangoFilterBackend)
    search_fields = ['hostname', 'sap_id', 'serial_number', 'order_id',
                     'model__name', 'model__vendor__name']
    ordering_fields = ['hostname', 'serial_number', 'sap_id', 'order_id',
                       'updated_at', 'created_at',
                       'model__name', 'model__vendor__name',
                       'model__type__name', 'state__name']
    ordering = ['hostname']
    filterset_class = AssetFilter

    def get_queryset(self):
        return Asset.objects.select_related(
            'model', 'model__vendor', 'model__type', 'state'
        ).all()

    def filter_queryset(self, queryset):
        for backend in self.filter_backends:
            queryset = backend().filter_queryset(self.request, queryset, self)
        return queryset

    @extend_schema(
        summary='Export assets as Excel (.xlsx)',
        responses={
            200: OpenApiResponse(
                response=OpenApiTypes.BINARY,
                description='Excel spreadsheet download',
            ),
        },
    )
    def get(self, request):
        qs = self.get_queryset()

        # Optional: filter by explicit list of IDs (comma-separated)
        ids_param = request.query_params.get('ids')
        if ids_param:
            id_list = [i.strip()
                       for i in ids_param.split(',') if i.strip().isdigit()]
            qs = qs.filter(pk__in=id_list)
        else:
            qs = self.filter_queryset(qs)

        # ── Build workbook ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Asset'
        ws.sheet_view.showGridLines = False

        # Header row
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(vertical='center', horizontal='left')

        ws.row_dimensions[1].height = 20

        # Data rows
        for row_idx, asset in enumerate(qs, start=2):
            fill = ROW_FILL_ODD if row_idx % 2 == 1 else ROW_FILL_EVEN
            for col_idx, (_, getter) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=getter(asset))
                cell.font = ROW_FONT
                cell.fill = fill
                cell.alignment = Alignment(
                    vertical='center', horizontal='left')

        # Auto column widths
        for col_idx in range(1, len(COLUMNS) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Freeze header
        ws.freeze_panes = 'A2'

        # ── Stream response ───────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = datetime.date.today().strftime('%Y-%m-%d')
        filename = f'asset_{date_str}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
